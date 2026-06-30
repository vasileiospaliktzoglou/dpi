from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, Any
import datetime as dt
import shutil

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from config import MEMORY_FILE, DATA_DIR

SHEETS = ["Runs", "ETF_Decisions", "Market_Summary", "Learning", "Feature_Store", "Execution_Journal"]
HEADERS = {
    "Runs": ["run_id", "timestamp", "version", "primary_etf", "primary_action"],
    "ETF_Decisions": ["run_id", "timestamp", "etf", "action", "live_price", "target_price", "gap_pct", "confidence", "target_touch_1d", "target_touch_5d", "fair_value", "expected_low", "expected_high", "better_price_1d", "better_price_2d", "better_price_3d", "reason"],
    "Market_Summary": ["run_id", "timestamp", "regime", "score", "summary", "drivers"],
    "Learning": ["run_id", "timestamp", "note"],
    "Feature_Store": ["run_id", "timestamp", "etf", "atr", "rsi", "trend", "market_regime", "market_score", "gap_pct", "fair_value_gap_pct", "expected_low", "expected_high", "better_price_3d", "decision"],
    "Execution_Journal": ["timestamp", "etf", "action", "execution_price", "shares", "amount_eur", "live_price", "suggested_limit", "fair_value", "expected_low", "expected_high", "better_price_1d", "better_price_2d", "better_price_3d", "confidence", "notes"],
}


def _safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        return "" if pd.isna(value) else float(value)
    if isinstance(value, (int, str, bool)):
        return value
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    if isinstance(value, (list, tuple, set)):
        return "; ".join(str(x) for x in value)
    if isinstance(value, dict):
        return "; ".join(f"{k}: {v}" for k, v in value.items())
    return str(value)


def _new_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEETS[0]
    for name in SHEETS[1:]:
        wb.create_sheet(name)
    for sheet, cols in HEADERS.items():
        wb[sheet].append(cols)
    wb.save(path)


def ensure_workbook() -> Path:
    Path(DATA_DIR).mkdir(parents=True, exist_ok=True)
    path = Path(MEMORY_FILE)
    if not path.exists():
        _new_workbook(path)
        return path
    try:
        wb = load_workbook(path)
        changed = False
        for sheet in SHEETS:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
                wb[sheet].append(HEADERS[sheet])
                changed = True
            elif wb[sheet].max_row == 0 or all(c.value is None for c in wb[sheet][1]):
                wb[sheet].append(HEADERS[sheet])
                changed = True
        if changed:
            wb.save(path)
    except (InvalidFileException, OSError, TypeError, ValueError):
        backup = path.with_suffix(f".broken_{dt.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
        try:
            shutil.copy2(path, backup)
        except Exception:
            pass
        _new_workbook(path)
    return path


def append_rows(sheet_name: str, rows: Iterable[dict]) -> None:
    path = ensure_workbook()
    rows = list(rows)
    if not rows:
        return
    try:
        wb = load_workbook(path)
    except Exception:
        _new_workbook(path)
        wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
        wb[sheet_name].append(HEADERS.get(sheet_name, list(rows[0].keys())))
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1] if c.value is not None]
    if not headers:
        headers = HEADERS.get(sheet_name, list(rows[0].keys()))
        ws.append(headers)
    for row in rows:
        ws.append([_safe_cell(row.get(h, "")) for h in headers])
    wb.save(path)


def save_run(version: str, market, decisions: Dict[str, Any], primary) -> None:
    now_dt = dt.datetime.now()
    now = now_dt.strftime("%Y-%m-%d %H:%M:%S")
    run_id = now_dt.strftime("%Y%m%d%H%M%S")
    append_rows("Runs", [{"run_id": run_id, "timestamp": now, "version": version, "primary_etf": primary.symbol, "primary_action": primary.action}])
    append_rows("Market_Summary", [{"run_id": run_id, "timestamp": now, "regime": market.regime, "score": market.score, "summary": market.one_sentence, "drivers": market.drivers}])
    append_rows("ETF_Decisions", [{
        "run_id": run_id, "timestamp": now, "etf": d.symbol, "action": d.action,
        "live_price": d.live_price, "target_price": d.target_price, "gap_pct": d.gap_pct,
        "confidence": d.confidence_label, "target_touch_1d": d.target_touch_1d,
        "target_touch_5d": d.target_touch_5d, "fair_value": d.fair_value,
        "expected_low": d.expected_low, "expected_high": d.expected_high,
        "better_price_1d": d.better_price_1d, "better_price_2d": d.better_price_2d,
        "better_price_3d": d.better_price_3d, "reason": d.reason,
    } for d in decisions.values()])
    append_rows("Feature_Store", [{
        "run_id": run_id, "timestamp": now, "etf": d.symbol, "atr": d.atr, "rsi": d.rsi,
        "trend": d.trend, "market_regime": market.regime, "market_score": market.score,
        "gap_pct": d.gap_pct, "fair_value_gap_pct": d.fair_value_gap_pct,
        "expected_low": d.expected_low, "expected_high": d.expected_high,
        "better_price_3d": d.better_price_3d, "decision": d.action,
    } for d in decisions.values()])


def append_execution(row: Dict[str, Any]) -> None:
    """Save one real-world execution or skip/cancel decision.

    This journal is intentionally simple: over time the rows can be compared with
    later market prices to improve the model's limit-price and patience rules.
    """
    append_rows("Execution_Journal", [row])
